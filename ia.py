# ia.py
# Modulo de inteligencia artificial
# Basado en: chatgpt.py, asistente.py, gemini.py del inge
#
# Requiere: pip install openai
# Variable de entorno: OPENAI_API_KEY

import os
import base64
import mimetypes
import zipfile
import re
from pathlib import Path
from openai import OpenAI

# Inicializar cliente OpenAI (igual que chatgpt.py del inge)
client = OpenAI()


# ── TIPOS DE ARCHIVO SOPORTADOS ──────────────────────────────────────────────

EXTENSIONES = {
    # Texto
    ".txt": "texto",  ".md": "texto",   ".log": "texto",
    ".csv": "texto",  ".json": "texto", ".xml": "texto",
    ".yaml": "texto", ".yml": "texto",  ".html": "texto",
    ".ini": "texto",  ".cfg": "texto",  ".toml": "texto",
    # Codigo fuente (40+ lenguajes)
    ".py": "codigo",    ".js": "codigo",  ".ts": "codigo",
    ".java": "codigo",  ".c": "codigo",   ".cpp": "codigo",
    ".cs": "codigo",    ".go": "codigo",  ".rs": "codigo",
    ".rb": "codigo",    ".php": "codigo", ".swift": "codigo",
    ".kt": "codigo",    ".r": "codigo",   ".sh": "codigo",
    ".sql": "codigo",   ".dart": "codigo",".lua": "codigo",
    ".hs": "codigo",    ".ex": "codigo",  ".scala": "codigo",
    # Imagen (incluye capturas de graficas matplotlib y OpenGL)
    ".jpg": "imagen",  ".jpeg": "imagen", ".png": "imagen",
    ".gif": "imagen",  ".bmp": "imagen",  ".webp": "imagen",
    # Audio
    ".mp3": "audio",  ".wav": "audio",  ".ogg": "audio",
    ".flac": "audio", ".m4a": "audio",  ".aac": "audio",
    # Video
    ".mp4": "video",  ".mkv": "video",  ".avi": "video",
    ".mov": "video",  ".webm": "video",
    # Documentos
    ".pdf": "documento",  ".docx": "documento",
    ".doc": "documento",  ".xlsx": "hoja",
    ".xls": "hoja",       ".odt": "documento",
}

LENGUAJES = {
    ".py": "Python",    ".js": "JavaScript", ".ts": "TypeScript",
    ".java": "Java",    ".c": "C",           ".cpp": "C++",
    ".cs": "C#",        ".go": "Go",         ".rs": "Rust",
    ".rb": "Ruby",      ".php": "PHP",       ".swift": "Swift",
    ".kt": "Kotlin",    ".r": "R",           ".sh": "Shell",
    ".sql": "SQL",      ".dart": "Dart",     ".lua": "Lua",
    ".scala": "Scala",  ".hs": "Haskell",    ".ex": "Elixir",
}


# ── FUNCION PRINCIPAL ─────────────────────────────────────────────────────────

def analizar(ruta_o_bytes, prompt, tipo_mime=None):
    """
    Analiza CUALQUIER tipo de archivo y responde la pregunta del usuario.
    Esta es la funcion principal de la libreria.

    Acepta:
        - Archivos de texto, codigo, imagen, audio, video, PDF, DOCX, XLSX
        - Cualquier idioma (la IA lo detecta automaticamente)
        - Bytes en memoria (para graficas matplotlib/OpenGL sin guardar a disco)

    Parametros:
        ruta_o_bytes -- ruta al archivo (str) O bytes en memoria
        prompt       -- pregunta o instruccion del usuario
        tipo_mime    -- solo necesario si se pasan bytes (ej: "image/png")

    Retorna:
        respuesta como string
    """
    # Caso 1: bytes en memoria (grafica matplotlib, frame OpenGL)
    if isinstance(ruta_o_bytes, bytes):
        if tipo_mime is None:
            tipo_mime = "image/png"
        return _consultar_imagen_bytes(ruta_o_bytes, tipo_mime, prompt)

    # Caso 2: archivo en disco
    ruta = str(ruta_o_bytes)
    ext  = Path(ruta).suffix.lower()
    tipo = EXTENSIONES.get(ext)

    if tipo is None:
        # Intentar por MIME
        mime, _ = mimetypes.guess_type(ruta)
        if mime:
            if mime.startswith("text/"):    tipo = "texto"
            elif mime.startswith("image/"): tipo = "imagen"
            elif mime.startswith("audio/"): tipo = "audio"
            elif mime.startswith("video/"): tipo = "video"

    if tipo is None:
        raise ValueError(
            f"Tipo de archivo no reconocido: '{ext}'\n"
            f"Soportados: texto, codigo, imagen, audio, video, PDF, DOCX, XLSX"
        )

    # Extraer contenido segun tipo
    if tipo in ("texto", "codigo"):
        return _consultar_texto_archivo(ruta, ext, prompt)
    elif tipo == "imagen":
        return _consultar_imagen_archivo(ruta, prompt)
    elif tipo in ("audio", "video"):
        return _consultar_audio_video(ruta, prompt)
    elif tipo == "documento":
        return _consultar_documento(ruta, prompt)
    elif tipo == "hoja":
        return _consultar_hoja(ruta, prompt)


# ── METODOS DE ATAJO ──────────────────────────────────────────────────────────

def preguntar(texto_o_ruta, pregunta):
    """Alias de analizar(). Mas natural para el usuario."""
    return analizar(texto_o_ruta, pregunta)


def resumir(ruta):
    """Resume el contenido de cualquier archivo."""
    return analizar(ruta, "Resume el contenido de este archivo de forma clara y concisa.")


def traducir(ruta, idioma_destino="español"):
    """Traduce el contenido de un archivo al idioma indicado."""
    return analizar(ruta, f"Traduce el contenido al {idioma_destino}. Mantén el formato.")


def revisar_codigo(ruta):
    """Hace code review de un archivo de codigo."""
    return analizar(
        ruta,
        "Haz un code review profesional. Identifica bugs, malas practicas "
        "y sugiere mejoras especificas con ejemplos."
    )


def describir_grafica(figura_matplotlib=None, ruta_imagen=None):
    """
    Describe e interpreta una grafica.
    Acepta figura matplotlib en memoria o imagen en disco.
    Basado en gemini.py del inge (vision de imagenes).
    """
    if figura_matplotlib is not None:
        from graficas import capturar_como_bytes
        datos = capturar_como_bytes(figura_matplotlib)
        return _consultar_imagen_bytes(datos, "image/png",
            "Analiza e interpreta esta grafica. Describe tendencias, valores "
            "importantes y que conclusiones se pueden sacar.")
    elif ruta_imagen is not None:
        return _consultar_imagen_archivo(ruta_imagen,
            "Analiza e interpreta esta grafica. Describe tendencias, valores "
            "importantes y que conclusiones se pueden sacar.")
    else:
        raise ValueError("Debes pasar figura_matplotlib o ruta_imagen.")


# ── VERIFICACION ──────────────────────────────────────────────────────────────

def verificar_conexion():
    """
    Verifica que la API key funcione.
    Basado en verificar.py del inge.
    """
    try:
        print("Verificando conexion con OpenAI...")
        print(f"API key: {client.api_key[:8]}...")
        r = client.responses.create(model="gpt-4o", input="di solo: ok")
        print(f"Conexion exitosa!")
        return True
    except Exception as e:
        print(f"Error de conexion: {e}")
        return False


# ── INTERNOS ──────────────────────────────────────────────────────────────────

def _consultar_texto_plano(prompt, system="Eres un asistente util. Responde claro y conciso."):
    """Consulta basica de texto. Identica a chatgpt.py / asistente.py del inge."""
    respuesta = client.responses.create(
        model="gpt-4o",
        input=[
            {
                "role": "developer",
                "content": system,
            },
            {
                "role": "user",
                "content": prompt,
            },
        ],
    )
    return respuesta.output_text


def _consultar_texto_archivo(ruta, ext, prompt):
    """Lee el archivo de texto o codigo y lo manda a la IA."""
    contenido = Path(ruta).read_text(encoding="utf-8", errors="replace")
    lenguaje  = LENGUAJES.get(ext, "")

    # Si es codigo, envolverlo en bloque para que la IA lo identifique bien
    if lenguaje:
        contenido = f"```{lenguaje.lower()}\n{contenido}\n```"

    nombre = Path(ruta).name
    mensaje = (
        f"Archivo: {nombre}"
        + (f" | Lenguaje: {lenguaje}" if lenguaje else "")
        + f"\n\nContenido:\n{contenido}\n\nInstruccion: {prompt}"
    )

    return _consultar_texto_plano(mensaje)


def _consultar_imagen_archivo(ruta, prompt):
    """
    Manda una imagen a la IA para que la analice.
    Basado en gemini.py del inge (vision de imagenes).
    """
    with open(ruta, "rb") as f:
        datos = f.read()

    mime, _ = mimetypes.guess_type(ruta)
    mime = mime or "image/jpeg"
    return _consultar_imagen_bytes(datos, mime, prompt)


def _consultar_imagen_bytes(datos, tipo_mime, prompt):
    """
    Version para imagenes ya en memoria (bytes).
    Para graficas matplotlib y frames OpenGL sin guardar a disco.
    """
    imagen_b64 = base64.b64encode(datos).decode("utf-8")

    respuesta = client.responses.create(
        model="gpt-4o",
        input=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_image",
                        "image_url": f"data:{tipo_mime};base64,{imagen_b64}",
                    },
                    {
                        "type": "input_text",
                        "text": prompt,
                    },
                ],
            }
        ],
    )
    return respuesta.output_text


def _consultar_audio_video(ruta, prompt):
    """
    Transcribe audio/video con speech_recognition y luego consulta a la IA.
    Basado en ejemplo2.py del inge.
    """
    import speech_recognition as sr

    recognizer = sr.Recognizer()
    ext = Path(ruta).suffix.lower()

    # Convertir a WAV si no lo es (speech_recognition solo acepta WAV nativamente)
    ruta_wav = ruta
    if ext != ".wav":
        try:
            from pydub import AudioSegment
            audio = AudioSegment.from_file(ruta)
            ruta_wav = ruta.replace(ext, "_temp.wav")
            audio.export(ruta_wav, format="wav")
        except ImportError:
            raise ImportError(
                "Para audio/video que no sea WAV instala pydub:\n"
                "  pip install pydub"
            )

    with sr.AudioFile(ruta_wav) as source:
        print(f"Transcribiendo: {Path(ruta).name}")
        audio_data = recognizer.record(source)

    try:
        transcripcion = recognizer.recognize_google(audio_data, language="es-ES")
        print(f"Transcripcion: {transcripcion}")
    except sr.UnknownValueError:
        transcripcion = "[No se pudo transcribir el audio]"
    except sr.RequestError as e:
        transcripcion = f"[Error al transcribir: {e}]"

    # Limpiar archivo temporal
    if ruta_wav != ruta and os.path.exists(ruta_wav):
        os.remove(ruta_wav)

    nombre = Path(ruta).name
    mensaje = (
        f"Archivo de audio/video: {nombre}\n\n"
        f"Transcripcion:\n{transcripcion}\n\n"
        f"Instruccion: {prompt}"
    )
    return _consultar_texto_plano(mensaje)


def _consultar_documento(ruta, prompt):
    """PDF o DOCX: extrae texto y consulta a la IA."""
    ext  = Path(ruta).suffix.lower()
    texto = None

    if ext == ".pdf":
        # Intentar PyMuPDF
        try:
            import fitz
            doc   = fitz.open(ruta)
            pages = [p.get_text().strip() for p in doc if p.get_text().strip()]
            doc.close()
            texto = "\n\n".join(f"[Pagina {i+1}]\n{t}" for i, t in enumerate(pages))
        except ImportError:
            pass

        # Fallback pdfplumber
        if texto is None:
            try:
                import pdfplumber
                with pdfplumber.open(ruta) as pdf:
                    pages = [p.extract_text() for p in pdf.pages if p.extract_text()]
                texto = "\n\n".join(pages)
            except ImportError:
                pass

    elif ext in (".docx", ".odt"):
        # Intentar python-docx
        try:
            from docx import Document
            doc   = Document(ruta)
            texto = "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
        except ImportError:
            pass

        # Fallback: extraer XML del ZIP interno con stdlib
        if texto is None:
            try:
                with zipfile.ZipFile(ruta) as z:
                    xml = z.read("word/document.xml").decode("utf-8", errors="replace")
                texto = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", xml)).strip()
            except Exception:
                pass

    if not texto:
        raise ValueError(
            f"No se pudo extraer texto de '{Path(ruta).name}'.\n"
            f"Instala: pip install PyMuPDF   (para PDF)\n"
            f"         pip install python-docx (para DOCX)"
        )

    nombre  = Path(ruta).name
    mensaje = (
        f"Documento: {nombre}\n\n"
        f"Contenido:\n{texto}\n\n"
        f"Instruccion: {prompt}"
    )
    return _consultar_texto_plano(mensaje)


def _consultar_hoja(ruta, prompt):
    """XLSX o CSV: convierte a texto tabular y consulta a la IA."""
    ext  = Path(ruta).suffix.lower()
    texto = None

    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb    = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            hojas = []
            for nombre_hoja in wb.sheetnames:
                ws   = wb[nombre_hoja]
                filas = [
                    "\t".join(str(c) if c is not None else "" for c in fila)
                    for fila in ws.iter_rows(values_only=True)
                ]
                hojas.append(f"[Hoja: {nombre_hoja}]\n" + "\n".join(filas))
            wb.close()
            texto = "\n\n".join(hojas)
        except ImportError:
            raise ImportError("Para XLSX instala: pip install openpyxl")

    elif ext == ".csv":
        import csv
        with open(ruta, newline="", encoding="utf-8", errors="replace") as f:
            filas = ["\t".join(fila) for fila in csv.reader(f)]
        texto = "\n".join(filas)

    nombre  = Path(ruta).name
    mensaje = (
        f"Archivo de datos: {nombre}\n\n"
        f"Contenido:\n{texto}\n\n"
        f"Instruccion: {prompt}"
    )
    return _consultar_texto_plano(mensaje)
