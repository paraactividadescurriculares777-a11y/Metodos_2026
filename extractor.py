# extractor.py
# Autora: Dana
# Convierte cualquier tipo de archivo en contenido listo para mandar a Claude.
#
# Estrategia por tipo:
#   texto/codigo    -> leer como string (stdlib)
#   imagen          -> base64 (stdlib)
#   audio/video     -> base64 (Claude los transcribe directamente)
#   PDF             -> extraer texto con PyMuPDF si esta, si no base64
#   DOCX            -> extraer texto con python-docx si esta,
#                      si no parsear el XML interno (stdlib pura)
#   XLSX            -> leer con openpyxl si esta, si no csv fallback
#   opengl          -> recibe bytes directamente (viene de Eddy)

import base64
import zipfile
import re
from pathlib import Path
from .exceptions import ExtractionError


def extraer(ruta: str, categoria: str, extra: str, cfg) -> dict:
    """
    Extrae el contenido de un archivo y lo prepara para Claude.

    Parametros:
        ruta      -- ruta al archivo
        categoria -- resultado de detector.detectar()
        extra     -- lenguaje de programacion si aplica
        cfg       -- Config de OmniSense (puede ser None en tests)

    Retorna:
        {
            "tipo":     "text" | "binario",
            "contenido": str (si texto) o bytes (si binario),
            "mime":     mime type del contenido,
            "meta":     dict con informacion adicional
        }
    """
    manejadores = {
        "text":        _extraer_texto,
        "code":        _extraer_codigo,
        "image":       _extraer_binario,
        "audio":       _extraer_binario,
        "video":       _extraer_binario,
        "document":    _extraer_documento,
        "spreadsheet": _extraer_hoja,
        "opengl":      _extraer_opengl,
    }

    fn = manejadores.get(categoria)
    if fn is None:
        raise ExtractionError(f"Sin extractor para categoria: '{categoria}'")

    return fn(ruta, extra, cfg)


def extraer_desde_bytes(datos: bytes, mime: str) -> dict:
    """
    Version especial para capturas OpenGL y frames en memoria.
    Eddy la llama desde opengl_capture.py con los bytes del frame.

    Parametros:
        datos -- bytes de la imagen (PNG recomendado)
        mime  -- mime type, normalmente "image/png"

    Retorna el mismo formato dict que extraer().
    """
    if not datos:
        raise ExtractionError("Los bytes estan vacios.")

    return {
        "tipo": "binario",
        "contenido": datos,
        "mime": mime,
        "meta": {
            "fuente": "opengl_frame",
            "size_kb": round(len(datos) / 1024, 1),
        },
    }


# --- Extractores internos ---

def _extraer_texto(ruta, extra, cfg):
    try:
        contenido = Path(ruta).read_text(encoding="utf-8", errors="replace")
        return {
            "tipo": "text",
            "contenido": contenido,
            "mime": "text/plain",
            "meta": {
                "chars": len(contenido),
                "lineas": contenido.count("\n") + 1,
            },
        }
    except Exception as e:
        raise ExtractionError(f"No se pudo leer el archivo de texto: {e}") from e


def _extraer_codigo(ruta, extra, cfg):
    try:
        contenido = Path(ruta).read_text(encoding="utf-8", errors="replace")
        lenguaje = extra or Path(ruta).suffix.lstrip(".")
        # Envolver en bloque de codigo para que Claude lo identifique bien
        contenido_formateado = f"```{lenguaje.lower()}\n{contenido}\n```"
        return {
            "tipo": "text",
            "contenido": contenido_formateado,
            "mime": "text/plain",
            "meta": {
                "lenguaje": lenguaje,
                "lineas": contenido.count("\n") + 1,
            },
        }
    except Exception as e:
        raise ExtractionError(f"No se pudo leer el archivo de codigo: {e}") from e


def _extraer_binario(ruta, extra, cfg):
    """Para imagen, audio y video: leer como bytes y mandar a Claude."""
    import mimetypes
    try:
        datos = Path(ruta).read_bytes()
        mime, _ = mimetypes.guess_type(ruta)
        mime = mime or "application/octet-stream"
        return {
            "tipo": "binario",
            "contenido": datos,
            "mime": mime,
            "meta": {"size_kb": round(len(datos) / 1024, 1)},
        }
    except Exception as e:
        raise ExtractionError(f"No se pudo leer el archivo binario: {e}") from e


def _extraer_opengl(ruta, extra, cfg):
    """Mismo que binario, pero con meta especifica de opengl."""
    resultado = _extraer_binario(ruta, extra, cfg)
    resultado["meta"]["fuente"] = "opengl_file"
    return resultado


def _extraer_documento(ruta, extra, cfg):
    """PDF o DOCX: primero intenta extraccion de texto, si no hay libs usa base64."""
    ext = Path(ruta).suffix.lower()
    if ext == ".pdf":
        return _extraer_pdf(ruta)
    elif ext in (".docx", ".doc", ".odt"):
        return _extraer_docx(ruta)
    else:
        # RTF u otros: leer como texto
        return _extraer_texto(ruta, extra, cfg)


def _extraer_pdf(ruta):
    # Opcion 1: PyMuPDF (rapido, liviano)
    try:
        import fitz
        doc = fitz.open(ruta)
        paginas = []
        for i, pagina in enumerate(doc):
            texto = pagina.get_text().strip()
            if texto:
                paginas.append(f"[Pagina {i + 1}]\n{texto}")
        doc.close()
        contenido = "\n\n".join(paginas)
        return {
            "tipo": "text",
            "contenido": contenido,
            "mime": "text/plain",
            "meta": {"paginas": len(paginas), "metodo": "pymupdf"},
        }
    except ImportError:
        pass

    # Opcion 2: pdfplumber
    try:
        import pdfplumber
        with pdfplumber.open(ruta) as pdf:
            paginas = []
            for i, pagina in enumerate(pdf.pages):
                texto = pagina.extract_text()
                if texto and texto.strip():
                    paginas.append(f"[Pagina {i + 1}]\n{texto.strip()}")
        contenido = "\n\n".join(paginas)
        return {
            "tipo": "text",
            "contenido": contenido,
            "mime": "text/plain",
            "meta": {"paginas": len(paginas), "metodo": "pdfplumber"},
        }
    except ImportError:
        pass

    # Fallback: mandar el PDF como binario, Claude lo lee directamente
    return _extraer_binario(ruta, "", None)


def _extraer_docx(ruta):
    # Opcion 1: python-docx
    try:
        from docx import Document
        doc = Document(ruta)
        parrafos = [p.text for p in doc.paragraphs if p.text.strip()]
        contenido = "\n\n".join(parrafos)
        return {
            "tipo": "text",
            "contenido": contenido,
            "mime": "text/plain",
            "meta": {"parrafos": len(parrafos), "metodo": "python-docx"},
        }
    except ImportError:
        pass

    # Opcion 2: DOCX es un ZIP, extraer el XML interno con stdlib
    try:
        with zipfile.ZipFile(ruta) as z:
            xml = z.read("word/document.xml").decode("utf-8", errors="replace")
        # Quitar todos los tags XML y dejar solo el texto
        texto = re.sub(r"<[^>]+>", " ", xml)
        texto = re.sub(r"\s+", " ", texto).strip()
        return {
            "tipo": "text",
            "contenido": texto,
            "mime": "text/plain",
            "meta": {"metodo": "xml_stdlib"},
        }
    except Exception:
        pass

    # Fallback: binario
    return _extraer_binario(ruta, "", None)


def _extraer_hoja(ruta, extra, cfg):
    """XLSX o CSV: convierte a texto tabular para que Claude analice los datos."""
    ext = Path(ruta).suffix.lower()

    # Opcion 1: openpyxl para XLSX
    if ext in (".xlsx", ".xls"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(ruta, read_only=True, data_only=True)
            hojas = []
            for nombre in wb.sheetnames:
                ws = wb[nombre]
                filas = []
                for fila in ws.iter_rows(values_only=True):
                    # Convertir cada celda a string, reemplazar None por ""
                    fila_texto = "\t".join(
                        str(c) if c is not None else "" for c in fila
                    )
                    filas.append(fila_texto)
                hojas.append(f"[Hoja: {nombre}]\n" + "\n".join(filas))
            wb.close()
            contenido = "\n\n".join(hojas)
            return {
                "tipo": "text",
                "contenido": contenido,
                "mime": "text/plain",
                "meta": {"hojas": len(hojas), "metodo": "openpyxl"},
            }
        except ImportError:
            pass

    # Opcion 2: CSV con stdlib
    if ext == ".csv":
        import csv
        try:
            with open(ruta, newline="", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                filas = ["\t".join(fila) for fila in reader]
            contenido = "\n".join(filas)
            return {
                "tipo": "text",
                "contenido": contenido,
                "mime": "text/plain",
                "meta": {"filas": len(filas), "metodo": "csv_stdlib"},
            }
        except Exception as e:
            raise ExtractionError(f"No se pudo leer el CSV: {e}") from e

    # Fallback: leer como texto
    return _extraer_texto(ruta, extra, cfg)
